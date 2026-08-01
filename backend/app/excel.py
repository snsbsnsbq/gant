import io
import re

from openpyxl import Workbook, load_workbook

from app.models import Task


class ExcelValidationError(ValueError):
    """Raised when an uploaded Excel file fails validation."""

# Header aliases (Russian primary, English fallbacks) -> canonical field.
HEADER_MAP = {
    "задача": "name",
    "task": "name",
    "name": "name",
    "название": "name",
    "описание": "description",
    "description": "description",
    "исполнитель": "assignee",
    "assignee": "assignee",
    "owner": "assignee",
    "длительность": "duration",
    "duration": "duration",
    "предшественники": "predecessors",
    "predecessors": "predecessors",
    "depends on": "predecessors",
}

EXPORT_HEADERS = [
    ("name", "Задача"),
    ("description", "Описание"),
    ("assignee", "Исполнитель"),
    ("duration", "Длительность"),
    ("predecessors", "Предшественники"),
]


def _split_predecessors(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[,;\n]+", str(value))
    return [p.strip() for p in parts if p.strip()]


def parse_excel(data: bytes) -> list[dict]:
    """Parse and validate an uploaded xlsx into raw task dicts.

    Predecessors are returned as *names* (``predecessor_names``); the caller
    resolves them to ids after tasks are created.

    Raises :class:`ExcelValidationError` with a human-readable message when the
    file is not a valid task import.
    """
    if not data:
        raise ExcelValidationError("Файл пустой. Загрузите Excel-файл с задачами.")

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except ExcelValidationError:
        raise
    except Exception:
        raise ExcelValidationError(
            "Не удалось прочитать файл. Убедитесь, что это корректный Excel-файл (.xlsx)."
        )

    ws = wb.active
    if ws is None:
        raise ExcelValidationError("В файле нет ни одного листа с данными.")

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ExcelValidationError("Файл пустой: не найдена строка заголовков.")

    columns: dict[int, str] = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key in HEADER_MAP:
            columns[idx] = HEADER_MAP[key]

    if "name" not in columns.values():
        raise ExcelValidationError(
            "Не найден обязательный столбец с названием задачи "
            "(например «Задача» или «Task»)."
        )

    tasks: list[dict] = []
    errors: list[str] = []
    seen_names: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip fully empty rows

        record = {"name": "", "description": "", "assignee": "", "duration": 1}
        predecessor_names: list[str] = []
        for idx, field in columns.items():
            value = row[idx] if idx < len(row) else None
            if field == "duration":
                if value is None or str(value).strip() == "":
                    record["duration"] = 1
                else:
                    try:
                        duration = int(float(value))
                    except (TypeError, ValueError):
                        errors.append(
                            f"строка {row_number}: длительность «{value}» не является числом"
                        )
                        continue
                    if duration < 0:
                        errors.append(
                            f"строка {row_number}: длительность не может быть отрицательной"
                        )
                        continue
                    record["duration"] = duration
            elif field == "predecessors":
                predecessor_names = _split_predecessors(value)
            else:
                record[field] = str(value).strip() if value is not None else ""

        if not record["name"]:
            errors.append(f"строка {row_number}: не заполнено название задачи")
            continue

        key = record["name"].strip().lower()
        if key in seen_names:
            errors.append(
                f"строка {row_number}: задача «{record['name']}» повторяется"
            )
            continue
        seen_names.add(key)

        record["predecessor_names"] = predecessor_names
        tasks.append(record)

    # Validate that predecessors reference tasks present in the same file.
    for record in tasks:
        for name in record["predecessor_names"]:
            if name.strip().lower() not in seen_names:
                errors.append(
                    f"задача «{record['name']}»: предшественник «{name}» не найден среди задач"
                )

    if not tasks and not errors:
        raise ExcelValidationError("В файле нет ни одной задачи для импорта.")

    if errors:
        preview = "; ".join(errors[:10])
        if len(errors) > 10:
            preview += f" и ещё {len(errors) - 10} ошибок"
        raise ExcelValidationError(f"Файл содержит ошибки: {preview}.")

    return tasks


def build_excel(tasks: list[Task]) -> bytes:
    """Serialize tasks back into an xlsx matching the input format."""
    id_to_name = {t.id: t.name for t in tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append([title for _, title in EXPORT_HEADERS])

    for task in tasks:
        predecessor_names = ", ".join(
            id_to_name.get(pid, "") for pid in task.predecessors if pid in id_to_name
        )
        ws.append(
            [task.name, task.description, task.assignee, task.duration, predecessor_names]
        )

    for column_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
